"""Tests for the export service (Excel output and HTML building)."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from app.services.exporter import (
    TutorSchedule,
    build_schedule_html,
    export_to_excel,
)


def _sample() -> list[TutorSchedule]:
    return [
        TutorSchedule(
            tutor_name="מאיה",
            cells={(2, 3): "דני\nמתמטיקה", (0, 1): "קבוצה א\nאנגלית"},
        )
    ]


def test_export_to_excel_creates_file(tmp_path: Path) -> None:
    out = tmp_path / "schedule.xlsx"
    export_to_excel(out, _sample(), "מערכת")
    assert out.exists()

    wb = load_workbook(out)
    assert "מאיה" in wb.sheetnames
    ws = wb["מאיה"]
    assert ws.sheet_view.rightToLeft
    # Header row of days present.
    assert ws.cell(row=2, column=1).value == "שעה"


def test_build_schedule_html_contains_data() -> None:
    html_text = build_schedule_html(_sample(), "מערכת")
    assert "dir='rtl'" in html_text
    assert "מאיה" in html_text
    assert "דני" in html_text
