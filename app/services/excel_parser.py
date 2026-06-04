"""Excel import parsing and validation (Pandas).

Parses a subject's ``.xlsx`` file into validated student records. This module
contains no Qt code so it can be unit-tested in isolation.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path

import pandas as pd

from app.config import (
    COL_CLASS,
    COL_LEVEL,
    COL_NAME,
    COL_UNITS,
    GRADE_TET,
    GRADE_YA,
    GRADE_YB,
    GRADE_YOD,
    LEVEL_MAX,
    LEVEL_MIN,
    REQUIRED_COLUMNS,
    UNITS_MAX,
    UNITS_MIN,
)

# Maps the bare Hebrew letter prefix to its canonical, punctuated grade label.
# Longer prefixes must be checked first (יא/יב before י).
_GRADE_PREFIXES: list[tuple[str, str]] = [
    ("יא", GRADE_YA),
    ("יב", GRADE_YB),
    ("ט", GRADE_TET),
    ("י", GRADE_YOD),
]

_CLASS_TOKEN_RE = re.compile(r"^([\u05d0-\u05ea]+)(\d+)$")


class ExcelImportError(Exception):
    """Raised when the file cannot be parsed at a structural level."""


@dataclass
class ParsedStudent:
    """A single validated student row ready to be persisted."""

    name: str
    grade: str
    class_number: int
    units: int
    study_level: int


@dataclass
class RowIssue:
    """A per-row validation problem reported back to the user."""

    row_number: int  # 1-based, matching the spreadsheet (header = row 1)
    message: str


@dataclass
class ImportResult:
    """Outcome of parsing a workbook: valid records plus any row issues."""

    students: list[ParsedStudent] = field(default_factory=list)
    issues: list[RowIssue] = field(default_factory=list)

    @property
    def has_issues(self) -> bool:
        return bool(self.issues)


def parse_class_token(raw: str) -> tuple[str, int]:
    """Parse a ``כיתה`` token (e.g. ``י"ב4``) into (canonical grade, class number).

    Handles geresh/gershayim variations such as ``ט'1``, ``י1``, ``יא2``.

    Raises:
        ValueError: if the token cannot be interpreted.
    """
    if raw is None:
        raise ValueError("ערך כיתה ריק")

    # Drop geresh ('), gershayim ("), common unicode variants and whitespace.
    cleaned = re.sub(r"[\s'\"\u05f3\u05f4\u2018\u2019\u201c\u201d]", "", str(raw))
    if not cleaned:
        raise ValueError("ערך כיתה ריק")

    match = _CLASS_TOKEN_RE.match(cleaned)
    if not match:
        raise ValueError(f"לא ניתן לפענח את הכיתה '{raw}'")

    letters, digits = match.group(1), match.group(2)
    for prefix, canonical in _GRADE_PREFIXES:
        if letters == prefix:
            return canonical, int(digits)

    raise ValueError(f"שכבה לא מזוהה בכיתה '{raw}'")


def _normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Strip whitespace from column headers for tolerant matching."""
    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]
    return df


def _coerce_int(value: object) -> int:
    """Coerce a spreadsheet cell to int, tolerating floats like ``5.0``."""
    if isinstance(value, bool):
        raise ValueError("ערך בוליאני אינו חוקי")
    if isinstance(value, (int,)):
        return int(value)
    if isinstance(value, float):
        if value.is_integer():
            return int(value)
        raise ValueError(f"ערך לא שלם '{value}'")
    text = str(value).strip()
    if not text:
        raise ValueError("ערך ריק")
    return int(float(text))


def parse_workbook(path: str | Path) -> ImportResult:
    """Read and validate a subject workbook into an :class:`ImportResult`.

    Raises:
        ExcelImportError: for file-level problems (unreadable file, missing
            required columns) that prevent any meaningful import.
    """
    path = Path(path)
    if not path.exists():
        raise ExcelImportError(f"הקובץ לא נמצא: {path}")

    try:
        df = pd.read_excel(path, dtype=object)
    except Exception as exc:  # noqa: BLE001 - surface any pandas/openpyxl error
        raise ExcelImportError(f"לא ניתן לקרוא את קובץ האקסל: {exc}") from exc

    df = _normalize_columns(df)

    missing = [c for c in REQUIRED_COLUMNS if c not in df.columns]
    if missing:
        raise ExcelImportError(
            "חסרות עמודות חובה בקובץ: " + ", ".join(missing)
        )

    result = ImportResult()
    for idx, row in df.iterrows():
        # Spreadsheet row number: header is row 1, first data row is row 2.
        row_number = int(idx) + 2 if isinstance(idx, int) else result_len(result) + 2

        name_raw = row.get(COL_NAME)
        name = "" if name_raw is None else str(name_raw).strip()
        if not name or name.lower() == "nan":
            # Skip fully blank rows silently; report partially blank rows.
            if all(_is_blank(row.get(c)) for c in REQUIRED_COLUMNS):
                continue
            result.issues.append(RowIssue(row_number, "חסר שם תלמיד"))
            continue

        try:
            grade, class_number = parse_class_token(row.get(COL_CLASS))
        except ValueError as exc:
            result.issues.append(RowIssue(row_number, f"{name}: {exc}"))
            continue

        try:
            units = _coerce_int(row.get(COL_UNITS))
        except ValueError:
            result.issues.append(
                RowIssue(row_number, f"{name}: יח\"ל לא תקין '{row.get(COL_UNITS)}'")
            )
            continue
        if not UNITS_MIN <= units <= UNITS_MAX:
            result.issues.append(
                RowIssue(row_number, f"{name}: יח\"ל חייב להיות בין {UNITS_MIN} ל-{UNITS_MAX}")
            )
            continue

        try:
            level = _coerce_int(row.get(COL_LEVEL))
        except ValueError:
            result.issues.append(
                RowIssue(row_number, f"{name}: רמת לימוד לא תקינה '{row.get(COL_LEVEL)}'")
            )
            continue
        if not LEVEL_MIN <= level <= LEVEL_MAX:
            result.issues.append(
                RowIssue(
                    row_number,
                    f"{name}: רמת לימוד חייבת להיות בין {LEVEL_MIN} ל-{LEVEL_MAX}",
                )
            )
            continue

        result.students.append(
            ParsedStudent(
                name=name,
                grade=grade,
                class_number=class_number,
                units=units,
                study_level=level,
            )
        )

    return result


def write_template(path: str | Path) -> Path:
    """Write a styled, RTL example workbook with the required import columns.

    The template contains the four required headers and a few valid example
    rows so the user can see the expected ``כיתה`` token format and value
    ranges. Returns the path written.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    path = Path(path)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "תלמידים"
    sheet.sheet_view.rightToLeft = True

    headers = list(REQUIRED_COLUMNS)  # [שם תלמיד, כיתה, יח"ל, רמת לימוד]
    examples = [
        ["דנה כהן", 'י"א2', 5, 4],
        ["יוסי לוי", "י1", 4, 3],
        ["מאיה ישראלי", "ט'3", 3, 5],
        ["אבי דוד", "יב2", 5, 4],
    ]

    sheet.append(headers)
    for row in examples:
        sheet.append(row)

    header_fill = PatternFill("solid", fgColor="2D6CDF")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    center = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="D0D7E5")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, _ in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border
        sheet.column_dimensions[cell.column_letter].width = 18

    for row_idx in range(2, len(examples) + 2):
        for col in range(1, len(headers) + 1):
            cell = sheet.cell(row=row_idx, column=col)
            cell.alignment = center
            cell.border = border

    sheet.freeze_panes = "A2"
    workbook.save(path)
    return path


def _is_blank(value: object) -> bool:
    if value is None:
        return True
    text = str(value).strip()
    return not text or text.lower() == "nan"


def result_len(result: ImportResult) -> int:
    """Fallback row counter used when the DataFrame index is non-integer."""
    return len(result.students) + len(result.issues)
