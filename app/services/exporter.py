"""Schedule export to formatted Excel (openpyxl) and PDF (Qt QPrinter)."""

from __future__ import annotations

import html
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.config import DAYS, HOURS


@dataclass
class TutorSchedule:
    """A single tutor's weekly schedule as a {(day, hour): label} mapping."""

    tutor_name: str
    cells: dict[tuple[int, int], str] = field(default_factory=dict)


def _sanitize_sheet_title(name: str) -> str:
    """Return a worksheet-safe title (<=31 chars, no illegal characters)."""
    for ch in "[]:*?/\\":
        name = name.replace(ch, " ")
    name = name.strip() or "חונכת"
    return name[:31]


def export_to_excel(
    path: str | Path, schedules: list[TutorSchedule], title: str
) -> Path:
    """Write the schedules to an .xlsx file, one RTL sheet per tutor."""
    path = Path(path)
    wb = Workbook()
    wb.remove(wb.active)

    header_fill = PatternFill("solid", fgColor="2D6CDF")
    header_font = Font(color="FFFFFF", bold=True)
    hour_fill = PatternFill("solid", fgColor="EEF2F9")
    thin = Side(style="thin", color="C9D2E3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    used_titles: set[str] = set()
    for schedule in schedules or [TutorSchedule(title)]:
        base = _sanitize_sheet_title(schedule.tutor_name)
        sheet_title, n = base, 2
        while sheet_title in used_titles:
            sheet_title = f"{base[:28]}_{n}"
            n += 1
        used_titles.add(sheet_title)

        ws = wb.create_sheet(sheet_title)
        ws.sheet_view.rightToLeft = True

        ws.cell(row=1, column=1, value=schedule.tutor_name)
        ws.cell(row=1, column=1).font = Font(bold=True, size=14)

        # Header row (row 2): empty corner + day names.
        ws.cell(row=2, column=1, value="שעה")
        for c, day in enumerate(DAYS, start=2):
            ws.cell(row=2, column=c, value=day)

        for r, hour in enumerate(HOURS, start=3):
            ws.cell(row=r, column=1, value=f"שעה {hour}")
            for c, _day in enumerate(DAYS, start=2):
                label = schedule.cells.get((c - 2, hour), "")
                ws.cell(row=r, column=c, value=label)

        # Styling.
        max_col = len(DAYS) + 1
        max_row = len(HOURS) + 2
        for c in range(1, max_col + 1):
            cell = ws.cell(row=2, column=c)
            cell.fill = header_fill
            cell.font = header_font
            ws.column_dimensions[get_column_letter(c)].width = 22
        for r in range(3, max_row + 1):
            ws.cell(row=r, column=1).fill = hour_fill
            ws.cell(row=r, column=1).font = Font(bold=True)
            ws.row_dimensions[r].height = 38
        for r in range(2, max_row + 1):
            for c in range(1, max_col + 1):
                cell = ws.cell(row=r, column=c)
                cell.border = border
                cell.alignment = center

    wb.save(path)
    return path


def build_schedule_html(schedules: list[TutorSchedule], title: str) -> str:
    """Build an RTL HTML document rendering each tutor's schedule as a table."""
    parts: list[str] = [
        "<html><head><meta charset='utf-8'></head>",
        "<body dir='rtl' style=\"font-family:'Segoe UI',Arial;\">",
        f"<h1 style='text-align:center'>{html.escape(title)}</h1>",
    ]
    for schedule in schedules:
        parts.append(
            f"<h2 style='color:#2d6cdf'>{html.escape(schedule.tutor_name)}</h2>"
        )
        parts.append(
            "<table dir='rtl' border='1' cellspacing='0' cellpadding='6' "
            "width='100%' style='border-collapse:collapse;'>"
        )
        parts.append("<tr style='background:#2d6cdf;color:#fff;'>")
        parts.append("<th>שעה</th>")
        for day in DAYS:
            parts.append(f"<th>{html.escape(day)}</th>")
        parts.append("</tr>")

        for hour in HOURS:
            parts.append("<tr>")
            parts.append(
                f"<td style='background:#eef2f9;font-weight:bold;'>שעה {hour}</td>"
            )
            for day_index in range(len(DAYS)):
                label = schedule.cells.get((day_index, hour), "")
                cell = html.escape(label).replace("\n", "<br>")
                parts.append(f"<td style='text-align:center'>{cell}</td>")
            parts.append("</tr>")
        parts.append("</table><br>")

    parts.append("</body></html>")
    return "".join(parts)


def export_to_pdf(
    path: str | Path, schedules: list[TutorSchedule], title: str
) -> Path:
    """Render the schedules to a PDF using Qt's printing stack (no extra deps)."""
    # Imported lazily so the Excel path has no hard Qt dependency.
    from PySide6.QtGui import QPageLayout, QPageSize, QTextDocument
    from PySide6.QtPrintSupport import QPrinter

    path = Path(path)
    printer = QPrinter(QPrinter.PrinterMode.HighResolution)
    printer.setOutputFormat(QPrinter.OutputFormat.PdfFormat)
    printer.setOutputFileName(str(path))
    printer.setPageOrientation(QPageLayout.Orientation.Landscape)
    printer.setPageSize(QPageSize(QPageSize.PageSizeId.A4))

    doc = QTextDocument()
    doc.setDefaultStyleSheet("td{border:1px solid #ccc;}")
    # RTL direction is carried by the HTML's dir='rtl' attributes.
    doc.setHtml(build_schedule_html(schedules, title))
    doc.print_(printer)
    return path
